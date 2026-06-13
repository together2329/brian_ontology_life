#!/usr/bin/env python3
"""Import Brian's historical review archive into ontology-friendly indexes."""

from __future__ import annotations

import csv
import datetime as dt
import hashlib
import json
import math
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import openpyxl
import xlrd
import yaml


REVIEW_ROOT = Path("/Users/brian/Desktop/_Review")
OUTPUT_ROOT = Path("life/imports/review_archive")
PATTERN_OUTPUT = Path("life/patterns/review_archive_pattern_candidates.yaml")

CANONICAL_SOURCES = [
    {
        "id": "review_2018_time_jira",
        "path": REVIEW_ROOT / "2018/User_ Brian Choi_2018-01-01_2018-12-31.xlsx",
        "kind": "tempo_xlsx",
        "source_type": "time_log",
        "year_hint": 2018,
    },
    {
        "id": "review_2019_time_jira",
        "path": REVIEW_ROOT / "2019/Report_2019-01-01_2019-12-31.xlsx",
        "kind": "tempo_xlsx",
        "source_type": "time_log",
        "year_hint": 2019,
    },
    {
        "id": "review_2020_time_jira",
        "path": REVIEW_ROOT / "2020/Report_2020-01-01_2020-11-25.xlsx",
        "kind": "tempo_xlsx",
        "source_type": "time_log",
        "year_hint": 2020,
    },
    {
        "id": "review_2021_clockify_work",
        "path": REVIEW_ROOT
        / "2021/Clockify_Detailed_Report_01_01_2021-12_31_2021 - 복사본.xlsx",
        "kind": "clockify_xlsx",
        "source_type": "time_log",
        "year_hint": 2021,
    },
    {
        "id": "review_2021_life_logged_time",
        "path": REVIEW_ROOT / "2021/Report_2021-05-01_2021-09-25_Logged_Time.csv",
        "kind": "life_logged_csv",
        "source_type": "time_log",
        "year_hint": 2021,
    },
    {
        "id": "review_2022_2023_life_logged_time",
        "path": REVIEW_ROOT / "2023/Report_01_Jan_22_09_Jul_23_Logged_Time.xls",
        "kind": "life_logged_xls",
        "source_type": "time_log",
        "year_hint": 2023,
    },
]

REVIEW_YEARS = ["2018", "2019", "2020", "2021", "2023"]
AREAS = ["MIND", "BODY", "CAREER", "FINANCE", "RELATION", "MAINTENANCE", "FUN"]

AREA_KEYWORDS = {
    "MIND": [
        "[MIND]",
        "MIND",
        "독서",
        "목표",
        "리뷰",
        "감사",
        "메모",
        "명상",
        "생각",
        "공부",
        "일기",
        "회고",
    ],
    "BODY": [
        "[BODY]",
        "BODY",
        "수면",
        "식사",
        "식단",
        "운동",
        "샤워",
        "몸",
        "컨디션",
        "산책",
    ],
    "CAREER": [
        "[CAREER]",
        "CAREER",
        "회사",
        "업무",
        "Plaform",
        "Platform",
        "CTS",
        "Deeper",
        "Tachy",
        "X330",
        "SRTRI",
        "라닉스",
        "하이비",
        "PCIe",
        "CPU",
        "DMA",
        "RTL",
        "spyglass",
        "timing",
        "clock",
        "sdc",
        "cdc",
        "bus",
        "support",
        "블로그",
        "개발",
        "설계",
        "공정",
        "layout",
        "synthesis",
    ],
    "FINANCE": ["FINANCIAL", "FINANCE", "재무", "투자", "돈", "가계부", "자산"],
    "RELATION": ["RELATION", "[REL]", "관계", "아름", "사랑", "대화", "가족", "친구", "사람", "통화"],
    "MAINTENANCE": ["정리", "유지", "설겆", "설거지", "청소", "집", "빨래", "관리"],
    "FUN": ["FUN", "Fun", "휴식", "유튜브", "게임", "영화", "즐거"],
}

EMOTION_KEYWORDS = {
    "happy": ["행복", "기쁘", "즐거", "좋", "설레", "만족"],
    "grateful": ["감사"],
    "calm": ["안정", "편안", "차분"],
    "tired": ["피곤", "지침", "힘듦", "졸림"],
    "anxious": ["불안", "걱정", "두렵"],
    "stressed": ["스트레스", "압박", "부담"],
    "unhappy": ["불행", "우울", "슬픔", "외롭"],
    "angry": ["화가", "짜증", "분노"],
}

OBSERVATION_KEYWORDS = {
    "sleep": ["수면"],
    "meal": ["식사", "식단"],
    "workout": ["운동", "헬스", "스쿼트", "푸쉬업", "산책"],
    "social_interaction": ["아름", "사랑", "대화", "가족", "친구", "사람", "통화"],
    "maintenance": ["정리", "유지", "설겆", "설거지", "청소", "빨래"],
    "fun": ["휴식", "유튜브", "게임", "영화", "즐거"],
    "finance_decision": ["투자", "돈", "재무", "가계부", "자산"],
}


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def sha256_text(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r\n", "\n").replace("\r", "\n").strip()
    return re.sub(r"\n{3,}", "\n\n", text)


def text_excerpt(text: str, limit: int = 240) -> str:
    text = clean_text(text)
    if len(text) <= limit:
        return text
    return text[: limit - 1].rstrip() + "…"


def hours_to_float(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)
    if isinstance(value, dt.timedelta):
        return value.total_seconds() / 3600
    if isinstance(value, dt.time):
        return value.hour + value.minute / 60 + value.second / 3600
    text = str(value).strip()
    if not text:
        return None
    if re.match(r"^\d+:\d\d(:\d\d)?$", text):
        parts = [float(p) for p in text.split(":")]
        if len(parts) == 3:
            return parts[0] + parts[1] / 60 + parts[2] / 3600
        return parts[0] + parts[1] / 60
    try:
        return float(text)
    except ValueError:
        return None


def parse_datetime(value: Any) -> Optional[dt.datetime]:
    if isinstance(value, dt.datetime):
        return value
    if isinstance(value, dt.date):
        return dt.datetime.combine(value, dt.time())
    text = str(value or "").strip()
    if not text:
        return None

    formats = [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
        "%m/%d/%Y %I:%M:%S %p",
        "%m/%d/%Y %I:%M %p",
        "%m/%d/%Y %H:%M:%S",
        "%m/%d/%Y %H:%M",
        "%m/%d/%Y",
    ]
    for fmt in formats:
        try:
            return dt.datetime.strptime(text, fmt)
        except ValueError:
            pass

    match = re.search(r"(\d{1,2}/[A-Za-z]{3}/\d{2}) at (\d{1,2}:\d{2})", text)
    if match:
        return dt.datetime.strptime(match.group(1) + " " + match.group(2), "%d/%b/%y %H:%M")
    return None


def combine_date_time(date_value: Any, time_value: Any) -> Optional[dt.datetime]:
    if isinstance(date_value, dt.datetime):
        date_part = date_value.date()
    elif isinstance(date_value, dt.date):
        date_part = date_value
    else:
        parsed = parse_datetime(date_value)
        date_part = parsed.date() if parsed else None

    if isinstance(time_value, dt.datetime):
        time_part = time_value.time()
    elif isinstance(time_value, dt.time):
        time_part = time_value
    else:
        text = str(time_value or "").strip()
        time_part = None
        for fmt in ["%I:%M:%S %p", "%I:%M %p", "%H:%M:%S", "%H:%M"]:
            try:
                time_part = dt.datetime.strptime(text, fmt).time()
                break
            except ValueError:
                pass

    if date_part and time_part:
        return dt.datetime.combine(date_part, time_part)
    if date_part:
        return dt.datetime.combine(date_part, dt.time())
    return None


def normalize_score(label: str, text: str) -> Optional[int]:
    numeric = re.search(label + r"\s*[:：]?\s*(\d{1,3})", text)
    if numeric:
        return int(numeric.group(1))
    level_match = re.search(label + r"\s*[:：]?\s*(높음|높다|보통|낮음|낮다)", text)
    if not level_match:
        return None
    return {"높음": 90, "높다": 90, "보통": 70, "낮음": 40, "낮다": 40}[level_match.group(1)]


def extract_metrics(text: str) -> Dict[str, int]:
    text = clean_text(text)
    metrics = {}
    energy = normalize_score("에너지", text)
    happiness = normalize_score("행복지수", text) or normalize_score("행복", text)
    focus = normalize_score("집중력", text)
    if energy is not None:
        metrics["energy"] = energy
    if happiness is not None:
        metrics["happiness"] = happiness
    if focus is not None:
        metrics["focus"] = focus
    return metrics


def extract_emotions(text: str) -> List[str]:
    hits = []
    for tag, words in EMOTION_KEYWORDS.items():
        if any(word in text for word in words):
            hits.append(tag)
    return hits


def infer_area(
    issue_key: str = "",
    title: str = "",
    description: str = "",
    task: str = "",
    activity: str = "",
    component: str = "",
) -> Tuple[str, str]:
    text = " ".join(clean_text(x) for x in [issue_key, title, description, task, activity, component])
    bracket = re.search(r"\[(MIND|BODY|CAREER|FINANCE|FINANCIAL|RELATION|REL|MAINTENANCE|FUN)\]", text, re.I)
    if bracket:
        value = bracket.group(1).upper()
        if value == "FINANCIAL":
            value = "FINANCE"
        if value == "REL":
            value = "RELATION"
        return value, "high"

    key = clean_text(issue_key).upper()
    if key.startswith("BODY"):
        return "BODY", "high"
    if key.startswith("MIND"):
        return "MIND", "high"
    if key.startswith("FUN"):
        return "FUN", "high"

    scores = {area: 0 for area in AREAS}
    lowered = text.lower()
    for area, words in AREA_KEYWORDS.items():
        for word in words:
            if word and word.lower() in lowered:
                scores[area] += 1

    best = max(scores, key=scores.get)
    if scores[best] >= 2:
        return best, "medium"
    if scores[best] == 1:
        return best, "low"
    return "MIND", "low"


def infer_observation_type(area: str, title: str, description: str) -> str:
    title_text = clean_text(title)
    combined_text = clean_text(title + "\n" + description)

    # Prefer title-level signals. Descriptions often contain incidental words
    # such as "잠시" or "먹다" that should not change the object type.
    for observation_type, words in OBSERVATION_KEYWORDS.items():
        if any(word in title_text for word in words):
            return observation_type

    if area == "CAREER":
        return "work_session"
    if area == "BODY":
        if any(word in combined_text for word in ["헬스", "스쿼트", "푸쉬업"]):
            return "workout"
        return "time_block"
    if area == "RELATION":
        return "social_interaction"
    if area == "MAINTENANCE":
        return "maintenance"
    if area == "FUN":
        return "fun"

    for observation_type, words in OBSERVATION_KEYWORDS.items():
        # Use description fallback only for lower-risk categories.
        if observation_type in {"finance_decision", "social_interaction", "maintenance", "fun"} and any(
            word in combined_text for word in words
        ):
            return observation_type
    if area == "FINANCE":
        return "finance_decision"
    return "time_block"


def make_record_id(source_id: str, row_number: int, when: Optional[dt.datetime]) -> str:
    date_part = when.strftime("%Y%m%d_%H%M") if when else "unknown_date"
    return f"hist_{date_part}_{source_id}_r{row_number}"


def read_tempo_xlsx(source: Dict[str, Any]) -> Iterable[Dict[str, Any]]:
    workbook = openpyxl.load_workbook(source["path"], read_only=True, data_only=True)
    worksheet = workbook["Worklogs"]
    rows = worksheet.iter_rows(values_only=True)
    header = [clean_text(x) for x in next(rows)]

    def index(name: str) -> Optional[int]:
        try:
            return header.index(name)
        except ValueError:
            return None

    indexes = {
        name: index(name)
        for name in ["Issue Key", "Issue summary", "Work date", "Work Description", "Hours", "Activity Name", "Component"]
    }

    def get(row: Tuple[Any, ...], name: str) -> Any:
        idx = indexes[name]
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    for row_number, row in enumerate(rows, start=2):
        if not any(clean_text(v) for v in row):
            continue
        yield {
            "source_id": source["id"],
            "source_path": str(source["path"]),
            "source_row": row_number,
            "source_type": source["source_type"],
            "issue_key": clean_text(get(row, "Issue Key")),
            "title": clean_text(get(row, "Issue summary")),
            "when": parse_datetime(get(row, "Work date")),
            "hours": hours_to_float(get(row, "Hours")),
            "description": clean_text(get(row, "Work Description")),
            "activity": clean_text(get(row, "Activity Name")),
            "component": clean_text(get(row, "Component")),
        }
    workbook.close()


def read_clockify_xlsx(source: Dict[str, Any]) -> Iterable[Dict[str, Any]]:
    workbook = openpyxl.load_workbook(source["path"], read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = worksheet.iter_rows(values_only=True)
    header = [clean_text(x) for x in next(rows)]
    indexes = {name: idx for idx, name in enumerate(header)}

    def get(row: Tuple[Any, ...], name: str) -> Any:
        idx = indexes.get(name)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    for row_number, row in enumerate(rows, start=2):
        if not any(clean_text(v) for v in row):
            continue
        yield {
            "source_id": source["id"],
            "source_path": str(source["path"]),
            "source_row": row_number,
            "source_type": source["source_type"],
            "issue_key": "",
            "title": clean_text(get(row, "Task")),
            "when": combine_date_time(get(row, "Start Date"), get(row, "Start Time")),
            "hours": hours_to_float(get(row, "Duration (decimal)") or get(row, "Duration (h)")),
            "description": clean_text(get(row, "Description")),
            "activity": "",
            "component": clean_text(get(row, "Project")),
        }
    workbook.close()


def read_life_logged_csv(source: Dict[str, Any]) -> Iterable[Dict[str, Any]]:
    with source["path"].open(newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row_number, row in enumerate(reader, start=2):
            description = clean_text(row.get("Worklog"))
            when = parse_datetime(row.get("Date"))
            if not description or not when:
                continue
            yield {
                "source_id": source["id"],
                "source_path": str(source["path"]),
                "source_row": row_number,
                "source_type": source["source_type"],
                "issue_key": clean_text(row.get("Key")),
                "title": clean_text(row.get("Issue")),
                "when": when,
                "hours": hours_to_float(row.get("Logged")),
                "description": description,
                "activity": "",
                "component": clean_text(row.get("Project")),
            }


def read_life_logged_xls(source: Dict[str, Any]) -> Iterable[Dict[str, Any]]:
    workbook = xlrd.open_workbook(source["path"])
    sheet = workbook.sheet_by_index(0)
    header = [clean_text(x) for x in sheet.row_values(0)]
    indexes = {name: idx for idx, name in enumerate(header)}

    def get(row: List[Any], name: str) -> Any:
        idx = indexes.get(name)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    for row_idx in range(1, sheet.nrows):
        row = sheet.row_values(row_idx)
        description = clean_text(get(row, "Worklog"))
        when = parse_datetime(get(row, "Date"))
        if not description or not when:
            continue
        yield {
            "source_id": source["id"],
            "source_path": str(source["path"]),
            "source_row": row_idx + 1,
            "source_type": source["source_type"],
            "issue_key": clean_text(get(row, "Key")),
            "title": clean_text(get(row, "Issue")),
            "when": when,
            "hours": hours_to_float(get(row, "Logged")),
            "description": description,
            "activity": "",
            "component": clean_text(get(row, "Project")),
        }


def source_rows(source: Dict[str, Any]) -> Iterable[Dict[str, Any]]:
    if source["kind"] == "tempo_xlsx":
        return read_tempo_xlsx(source)
    if source["kind"] == "clockify_xlsx":
        return read_clockify_xlsx(source)
    if source["kind"] == "life_logged_csv":
        return read_life_logged_csv(source)
    if source["kind"] == "life_logged_xls":
        return read_life_logged_xls(source)
    raise ValueError(f"Unsupported source kind: {source['kind']}")


def normalize_row(row: Dict[str, Any]) -> Dict[str, Any]:
    title = row["title"] or "Untitled historical record"
    description = row["description"]
    area, area_confidence = infer_area(
        issue_key=row["issue_key"],
        title=title,
        description=description,
        activity=row.get("activity", ""),
        component=row.get("component", ""),
    )
    observation_type = infer_observation_type(area, title, description)
    when = row["when"]
    duration_minutes = round((row["hours"] or 0) * 60, 2) if row["hours"] is not None else None
    raw_text = clean_text(description)
    normalized = {
        "id": make_record_id(row["source_id"], row["source_row"], when),
        "source_ref": row["source_id"],
        "source_path": row["source_path"],
        "source_row": row["source_row"],
        "source_type": row["source_type"],
        "original_date": when.date().isoformat() if when else None,
        "start_time": when.strftime("%H:%M") if when else None,
        "area": area,
        "area_confidence": area_confidence,
        "observation_type": observation_type,
        "title": title,
        "source_key": row["issue_key"] or None,
        "duration_minutes": duration_minutes,
        "metrics": extract_metrics(raw_text),
        "emotion_tags": extract_emotions(raw_text + "\n" + title),
        "raw_text_sha256": sha256_text(raw_text) if raw_text else None,
        "raw_text_excerpt": text_excerpt(raw_text),
        "needs_user_review": area_confidence == "low",
    }
    if row.get("component"):
        normalized["legacy_component"] = row["component"]
    if row.get("activity"):
        normalized["legacy_activity"] = row["activity"]
    return normalized


def inventory_files() -> List[Dict[str, Any]]:
    canonical_paths = {source["path"].resolve(): source["id"] for source in CANONICAL_SOURCES}
    files = []
    for year in REVIEW_YEARS:
        root = REVIEW_ROOT / year
        for path in sorted(root.rglob("*")):
            if not path.is_file():
                continue
            if path.name == ".DS_Store" or path.name.startswith("~$"):
                role = "ignored_system_or_temp"
            elif path.resolve() in canonical_paths:
                role = "canonical"
            elif "backup" in [part.lower() for part in path.parts]:
                role = "backup_or_duplicate"
            else:
                role = "supplementary"
            files.append(
                {
                    "path": str(path),
                    "name": path.name,
                    "extension": path.suffix.lower().lstrip(".") or "no_ext",
                    "size_bytes": path.stat().st_size,
                    "sha256": sha256_file(path) if role != "ignored_system_or_temp" else None,
                    "role": role,
                    "canonical_source_ref": canonical_paths.get(path.resolve()),
                }
            )
    return files


def summarize_records(records: List[Dict[str, Any]], manifest_files: List[Dict[str, Any]]) -> Dict[str, Any]:
    by_year = Counter(record["original_date"][:4] for record in records if record["original_date"])
    by_area_records = Counter(record["area"] for record in records)
    by_area_minutes = defaultdict(float)
    by_source_records = Counter(record["source_ref"] for record in records)
    by_source_minutes = defaultdict(float)
    metric_counts = Counter()
    emotion_counts = Counter()
    observation_counts = Counter(record["observation_type"] for record in records)
    title_minutes = defaultdict(float)
    title_counts = Counter()

    for record in records:
        minutes = record["duration_minutes"] or 0
        by_area_minutes[record["area"]] += minutes
        by_source_minutes[record["source_ref"]] += minutes
        for metric in record["metrics"]:
            metric_counts[metric] += 1
        for tag in record["emotion_tags"]:
            emotion_counts[tag] += 1
        title_key = (record["area"], record["title"])
        title_minutes[title_key] += minutes
        title_counts[title_key] += 1

    top_titles = []
    for (area, title), minutes in sorted(title_minutes.items(), key=lambda item: item[1], reverse=True)[:40]:
        top_titles.append(
            {
                "area": area,
                "title": title,
                "records": title_counts[(area, title)],
                "hours": round(minutes / 60, 2),
            }
        )

    source_summary = []
    for source in CANONICAL_SOURCES:
        source_summary.append(
            {
                "id": source["id"],
                "path": str(source["path"]),
                "kind": source["kind"],
                "records": by_source_records[source["id"]],
                "hours": round(by_source_minutes[source["id"]] / 60, 2),
            }
        )

    return {
        "schema_version": 1,
        "generated_at": dt.datetime.now().replace(microsecond=0).isoformat(),
        "import_scope": {
            "roots": [str(REVIEW_ROOT / year) for year in REVIEW_YEARS],
            "file_count": len(manifest_files),
            "canonical_source_count": len(CANONICAL_SOURCES),
            "raw_policy": "raw files are referenced by path and sha256; full raw prose is not copied into this repo",
        },
        "records": {
            "normalized_record_count": len(records),
            "date_range": {
                "start": min(record["original_date"] for record in records if record["original_date"]),
                "end": max(record["original_date"] for record in records if record["original_date"]),
            },
            "by_year": dict(sorted(by_year.items())),
            "by_area_records": dict(by_area_records),
            "by_area_hours": {area: round(minutes / 60, 2) for area, minutes in sorted(by_area_minutes.items())},
            "by_observation_type": dict(observation_counts),
            "extracted_metric_counts": dict(metric_counts),
            "emotion_tag_counts": dict(emotion_counts),
        },
        "canonical_sources": source_summary,
        "top_legacy_titles_by_time": top_titles,
        "ontology_mapping": {
            "historical row": "ParsedObservation candidate",
            "date + start_time + duration_minutes": "TimeBlock candidate",
            "area": "Area link",
            "title/source_key": "legacy Task/Habit/Issue candidate",
            "metrics.energy/happiness/focus": "MoodEntry or Observation candidate",
            "sleep rows": "SleepRecord candidate",
            "meal rows": "Meal candidate",
            "workout rows": "WorkoutSession candidate",
            "relationship rows": "SocialInteraction candidate",
            "maintenance rows": "Maintenance Task candidate",
            "finance rows": "Finance note/decision candidate",
        },
    }


def daily_summaries(records: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    grouped: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for record in records:
        if record["original_date"]:
            grouped[record["original_date"]].append(record)

    summaries = []
    for date, items in sorted(grouped.items()):
        area_minutes = defaultdict(float)
        metric_values = defaultdict(list)
        emotion_counts = Counter()
        observation_counts = Counter()
        for item in items:
            area_minutes[item["area"]] += item["duration_minutes"] or 0
            for metric, value in item["metrics"].items():
                metric_values[metric].append(value)
            emotion_counts.update(item["emotion_tags"])
            observation_counts[item["observation_type"]] += 1
        summaries.append(
            {
                "date": date,
                "record_count": len(items),
                "area_minutes": {area: round(minutes, 2) for area, minutes in sorted(area_minutes.items())},
                "metric_averages": {
                    metric: round(sum(values) / len(values), 2) for metric, values in sorted(metric_values.items())
                },
                "emotion_tag_counts": dict(emotion_counts),
                "observation_counts": dict(observation_counts),
            }
        )
    return summaries


def pattern_candidates(records: List[Dict[str, Any]]) -> Dict[str, Any]:
    def matches(*needles: str, area: Optional[str] = None) -> List[Dict[str, Any]]:
        found = []
        for record in records:
            text = (record["title"] + "\n" + record["raw_text_excerpt"]).lower()
            if area and record["area"] != area:
                continue
            if all(needle.lower() in text for needle in needles):
                found.append(record)
        return found

    def evidence(records_: List[Dict[str, Any]], limit: int = 6) -> List[Dict[str, Any]]:
        out = []
        for record in records_[:limit]:
            out.append(
                {
                    "record_ref": record["id"],
                    "date": record["original_date"],
                    "area": record["area"],
                    "title": record["title"],
                    "source_ref": record["source_ref"],
                    "source_row": record["source_row"],
                }
            )
        return out

    sleep_energy = [
        record
        for record in records
        if record["observation_type"] == "sleep" and "energy" in record["metrics"] and record["duration_minutes"]
    ]
    gratitude_positive = [
        record
        for record in records
        if "감사" in (record["title"] + record["raw_text_excerpt"])
        and ("happy" in record["emotion_tags"] or "grateful" in record["emotion_tags"])
    ]
    relationship_areum = [
        record
        for record in records
        if "아름" in (record["title"] + record["raw_text_excerpt"])
        and ("happy" in record["emotion_tags"] or "happiness" in record["metrics"])
    ]
    maintenance_positive = [
        record
        for record in records
        if record["area"] == "MAINTENANCE"
        and record["observation_type"] == "maintenance"
        and ("happy" in record["emotion_tags"] or "calm" in record["emotion_tags"])
    ]
    goal_review_stability = matches("목표", "안정")
    career_energy = [
        record for record in records if record["area"] == "CAREER" and "energy" in record["metrics"]
    ]
    passive_fun_review = [
        record
        for record in records
        if record["area"] == "FUN"
        and any(word in (record["title"] + record["raw_text_excerpt"]) for word in ["유튜브", "휴식", "불행"])
    ]

    candidates = [
        {
            "id": "pattern_candidate_legacy_sleep_energy_tracking",
            "statement": "Historical records frequently connect sleep duration with energy notes.",
            "confidence": "medium",
            "needs_user_review": True,
            "evidence_summary": {
                "matching_records": len(sleep_energy),
                "total_sleep_hours": round(sum((r["duration_minutes"] or 0) for r in sleep_energy) / 60, 2),
            },
            "evidence": evidence(sleep_energy),
        },
        {
            "id": "pattern_candidate_legacy_gratitude_positive_affect",
            "statement": "Gratitude records often appear with positive emotional language.",
            "confidence": "medium",
            "needs_user_review": True,
            "evidence_summary": {"matching_records": len(gratitude_positive)},
            "evidence": evidence(gratitude_positive),
        },
        {
            "id": "pattern_candidate_legacy_relationship_areum_happiness",
            "statement": "Records involving '아름' frequently appear with happiness or positive affect.",
            "confidence": "medium",
            "needs_user_review": True,
            "evidence_summary": {"matching_records": len(relationship_areum)},
            "evidence": evidence(relationship_areum),
        },
        {
            "id": "pattern_candidate_legacy_maintenance_mood_boost",
            "statement": "Cleaning, organizing, and maintenance records sometimes mention improved mood or calm.",
            "confidence": "low",
            "needs_user_review": True,
            "evidence_summary": {"matching_records": len(maintenance_positive)},
            "evidence": evidence(maintenance_positive),
        },
        {
            "id": "pattern_candidate_legacy_goal_review_stability",
            "statement": "Goal-setting and review records sometimes explicitly mention stability or feeling settled.",
            "confidence": "low",
            "needs_user_review": True,
            "evidence_summary": {"matching_records": len(goal_review_stability)},
            "evidence": evidence(goal_review_stability),
        },
        {
            "id": "pattern_candidate_legacy_career_energy_tracking",
            "statement": "Career work sessions include enough energy notes to support future high-energy scheduling analysis.",
            "confidence": "medium",
            "needs_user_review": False,
            "evidence_summary": {"matching_records": len(career_energy)},
            "evidence": evidence(career_energy),
        },
        {
            "id": "pattern_candidate_legacy_fun_restoration_vs_avoidance",
            "statement": "FUN records include both rest/happiness language and '불행' language, so restorative fun and avoidance should be separated.",
            "confidence": "low",
            "needs_user_review": True,
            "evidence_summary": {"matching_records": len(passive_fun_review)},
            "evidence": evidence(passive_fun_review),
        },
    ]

    return {
        "schema_version": 1,
        "generated_at": dt.datetime.now().replace(microsecond=0).isoformat(),
        "source": "life/imports/review_archive/normalized_records.jsonl",
        "causality_policy": "candidate patterns only; do not treat as confirmed identity or causality without Brian review",
        "pattern_candidates": candidates,
    }


def yaml_dump(path: Path, data: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        yaml.safe_dump(data, allow_unicode=True, sort_keys=False, width=120),
        encoding="utf-8",
    )


def jsonl_dump(path: Path, rows: Iterable[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        for row in rows:
            f.write(json.dumps(row, ensure_ascii=False, sort_keys=True) + "\n")


def main() -> None:
    OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)
    manifest_files = inventory_files()

    records = []
    for source in CANONICAL_SOURCES:
        for row in source_rows(source):
            records.append(normalize_row(row))

    records.sort(key=lambda item: (item["original_date"] or "", item["start_time"] or "", item["id"]))
    daily = daily_summaries(records)

    manifest = {
        "schema_version": 1,
        "generated_at": dt.datetime.now().replace(microsecond=0).isoformat(),
        "review_roots": [str(REVIEW_ROOT / year) for year in REVIEW_YEARS],
        "canonical_sources": [
            {
                "id": source["id"],
                "path": str(source["path"]),
                "kind": source["kind"],
                "source_type": source["source_type"],
                "year_hint": source["year_hint"],
                "sha256": sha256_file(source["path"]),
            }
            for source in CANONICAL_SOURCES
        ],
        "files": manifest_files,
    }

    yaml_dump(OUTPUT_ROOT / "manifest.yaml", manifest)
    yaml_dump(OUTPUT_ROOT / "summary.yaml", summarize_records(records, manifest_files))
    jsonl_dump(OUTPUT_ROOT / "normalized_records.jsonl", records)
    jsonl_dump(OUTPUT_ROOT / "daily_summary.jsonl", daily)
    yaml_dump(PATTERN_OUTPUT, pattern_candidates(records))

    print(f"normalized_records={len(records)}")
    print(f"daily_summaries={len(daily)}")
    print(f"manifest_files={len(manifest_files)}")


if __name__ == "__main__":
    main()
